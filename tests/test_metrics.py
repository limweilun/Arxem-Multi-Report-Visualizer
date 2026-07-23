from __future__ import annotations

import math
import unittest
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from app import (
    INITIAL_EQUITY,
    TRADING_DAYS_PER_YEAR,
    build_daily_returns,
    build_drawdown_curve,
    build_equity_curve,
    build_equity_overlay_table,
    compute_risk_metrics,
    extract_realized_deals,
    make_download_workbook,
    max_drawdown_duration,
    parse_report,
)


def deal(
    time: str,
    deal_id: int,
    position: int | None,
    direction: str,
    profit: float = 0.0,
    commission: float = 0.0,
    fee: float = 0.0,
    swap: float = 0.0,
    volume: float = 1.0,
) -> dict:
    return {
        "Time": pd.Timestamp(time),
        "Deal": str(deal_id),
        "Position": str(position) if position is not None else None,
        "Order": None,
        "Symbol": "TEST",
        "Type": "buy",
        "Direction": direction,
        "Volume": volume,
        "Profit": profit,
        "Commission": commission,
        "Fee": fee,
        "Swap": swap,
        "NetPnl": profit + commission + fee + swap,
    }


class UploadedWorkbook:
    def __init__(self, payload: bytes, name: str = "synthetic.xlsx") -> None:
        self._payload = payload
        self.name = name

    def getvalue(self) -> bytes:
        return self._payload


class MetricCalculationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.deals = pd.DataFrame(
            [
                deal("2024-01-02 09:00", 1, 101, "in", commission=-2.0),
                deal("2024-01-02 16:00", 2, 101, "out", profit=102.0, commission=-2.0),
                deal("2024-01-03 09:00", 3, 102, "in", commission=-2.0),
                deal("2024-01-04 16:00", 4, 102, "out", profit=-96.0, commission=-2.0),
                deal("2024-01-05 09:00", 5, 103, "in", commission=-2.0),
                deal("2024-01-08 16:00", 6, 103, "out", profit=202.0, commission=-2.0),
            ]
        )

    def test_completed_positions_include_entry_costs_without_extra_trades(self) -> None:
        realized = extract_realized_deals(self.deals)

        self.assertEqual(len(realized), 3)
        self.assertEqual(realized["NetPnl"].tolist(), [98.0, -100.0, 198.0])
        self.assertAlmostEqual(realized["NetPnl"].sum(), self.deals["NetPnl"].sum())

    def test_every_metric_reconciles_to_fixed_balance_definitions(self) -> None:
        realized = extract_realized_deals(self.deals)
        equity = build_equity_curve(self.deals)
        drawdown = build_drawdown_curve(equity)
        daily = build_daily_returns(realized)
        metrics = compute_risk_metrics(
            realized,
            drawdown,
            daily,
            total_costs=-12.0,
        )

        expected_returns = pd.Series([0.00098, 0.0, -0.001, 0.0, 0.00198])
        expected_sharpe = (
            expected_returns.mean()
            / expected_returns.std(ddof=1)
            * math.sqrt(TRADING_DAYS_PER_YEAR)
        )
        expected_downside_deviation = math.sqrt(
            expected_returns.clip(upper=0).pow(2).mean()
        )
        expected_sortino = (
            expected_returns.mean()
            / expected_downside_deviation
            * math.sqrt(TRADING_DAYS_PER_YEAR)
        )
        expected_max_dd_pct = 100.0 / 100098.0

        self.assertAlmostEqual(metrics["Total Net Profit"], 196.0)
        self.assertAlmostEqual(metrics["Gross Profit"], 296.0)
        self.assertAlmostEqual(metrics["Gross Loss"], -100.0)
        self.assertAlmostEqual(metrics["Costs / Swap"], -12.0)
        self.assertAlmostEqual(metrics["Profit Factor"], 2.96)
        self.assertEqual(metrics["Total Trades"], 3)
        self.assertAlmostEqual(metrics["Win Rate"], 200.0 / 3.0)
        self.assertAlmostEqual(metrics["Expectancy per Trade"], 196.0 / 3.0)
        self.assertAlmostEqual(metrics["Average Win"], 148.0)
        self.assertAlmostEqual(metrics["Average Loss"], -100.0)
        self.assertAlmostEqual(metrics["Best Trade"], 198.0)
        self.assertAlmostEqual(metrics["Worst Trade"], -100.0)
        self.assertAlmostEqual(metrics["Largest Single Loss"], -100.0)
        self.assertAlmostEqual(metrics["Best Day"], 198.0)
        self.assertAlmostEqual(metrics["Worst Day"], -100.0)
        self.assertAlmostEqual(metrics["Balance Drawdown Maximal Value"], 100.0)
        self.assertAlmostEqual(metrics["Max Drawdown (%)"], expected_max_dd_pct)
        self.assertAlmostEqual(metrics["Recovery Factor"], 1.96)
        self.assertAlmostEqual(metrics["Return on $100k"], 0.00196)
        self.assertAlmostEqual(metrics["Annualized Return"], expected_returns.mean() * 252)
        self.assertAlmostEqual(metrics["Sharpe Ratio"], expected_sharpe)
        self.assertAlmostEqual(metrics["Sortino Ratio"], expected_sortino)
        self.assertAlmostEqual(
            metrics["Calmar Ratio"],
            metrics["Annualized Return"] / expected_max_dd_pct,
        )
        self.assertEqual(metrics["Time to Recovery"], "6d 0h")
        self.assertAlmostEqual(metrics["Time to Recovery Days"], 6.0)

    def test_daily_returns_include_zero_business_days(self) -> None:
        daily = build_daily_returns(extract_realized_deals(self.deals))

        self.assertEqual(
            daily["Date"].dt.strftime("%Y-%m-%d").tolist(),
            [
                "2024-01-02",
                "2024-01-03",
                "2024-01-04",
                "2024-01-05",
                "2024-01-08",
            ],
        )
        self.assertEqual(daily["NetPnl"].tolist(), [98.0, 0.0, -100.0, 0.0, 198.0])

    def test_positionless_export_allocates_entry_costs_without_counting_them(self) -> None:
        deals = self.deals.copy()
        deals["Position"] = None

        realized = extract_realized_deals(deals)

        self.assertEqual(len(realized), 3)
        self.assertAlmostEqual(realized["NetPnl"].sum(), deals["NetPnl"].sum())
        self.assertEqual(realized["NetPnl"].tolist(), [98.0, -100.0, 198.0])

    def test_same_timestamp_trades_do_not_create_phantom_drawdown(self) -> None:
        deals = pd.DataFrame(
            [
                deal("2024-01-02 09:00", 1, 101, "in"),
                deal("2024-01-02 16:00", 2, 101, "out", profit=-100.0),
                deal("2024-01-02 09:00", 3, 102, "in"),
                deal("2024-01-02 16:00", 4, 102, "out", profit=200.0),
            ]
        )

        equity = build_equity_curve(deals)
        drawdown = build_drawdown_curve(equity)

        self.assertEqual(equity["Equity"].tolist(), [INITIAL_EQUITY, 100100.0])
        self.assertEqual(drawdown["Drawdown"].max(), 0.0)

    def test_unrecovered_drawdown_duration_runs_from_prior_peak(self) -> None:
        curve = pd.DataFrame(
            {
                "Time": pd.to_datetime(
                    ["2024-01-01", "2024-01-05", "2024-01-12"]
                ),
                "Equity": [100000.0, 99000.0, 99500.0],
            }
        )
        curve = build_drawdown_curve(curve)

        self.assertEqual(max_drawdown_duration(curve).days, 11)

    def test_overlay_starts_each_report_at_initial_equity(self) -> None:
        first = pd.DataFrame(
            {
                "Time": pd.to_datetime(["2024-01-01", "2024-01-02"]),
                "Equity": [100000.0, 100100.0],
            }
        )
        second = pd.DataFrame(
            {
                "Time": pd.to_datetime(["2024-01-03", "2024-01-04"]),
                "Equity": [100000.0, 99900.0],
            }
        )
        overlay = build_equity_overlay_table(
            [
                {"name": "First", "equity": first},
                {"name": "Second", "equity": second},
            ]
        )

        self.assertEqual(overlay.loc[0, "Second"], INITIAL_EQUITY)
        self.assertEqual(overlay.iloc[-1]["First"], 100100.0)

    def test_parser_ignores_reported_metrics_and_export_is_auditable(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Deals"])
        sheet.append(
            [
                "Time",
                "Deal",
                "Position",
                "Symbol",
                "Type",
                "Direction",
                "Volume",
                "Commission",
                "Fee",
                "Swap",
                "Profit",
            ]
        )
        sheet.append(
            [
                pd.Timestamp("2024-01-02 09:00").to_pydatetime(),
                1,
                101,
                "TEST",
                "buy",
                "in",
                1.0,
                -2.0,
                0.0,
                0.0,
                0.0,
            ]
        )
        sheet.append(
            [
                pd.Timestamp("2024-01-03 16:00").to_pydatetime(),
                2,
                101,
                "TEST",
                "sell",
                "out",
                1.0,
                -2.0,
                0.0,
                0.0,
                104.0,
            ]
        )
        sheet.append(["Orders"])
        sheet.append(["Total Net Profit", 999999.0])
        sheet.append(["Sharpe Ratio", 0.01])
        buffer = BytesIO()
        workbook.save(buffer)

        report = parse_report(UploadedWorkbook(buffer.getvalue()))

        self.assertEqual(len(report["deals"]), 2)
        self.assertEqual(len(report["realized_deals"]), 1)
        self.assertAlmostEqual(report["metrics"]["Total Net Profit"], 100.0)
        self.assertNotEqual(report["metrics"]["Total Net Profit"], 999999.0)
        self.assertNotEqual(report["metrics"]["Sharpe Ratio"], 0.01)
        self.assertEqual(report["reported_metrics"]["Sharpe Ratio"], 0.01)

        exported = make_download_workbook([report])
        exported_workbook = load_workbook(BytesIO(exported), data_only=False)
        self.assertEqual(
            exported_workbook.sheetnames,
            [
                "Summary",
                "Equity Overlay",
                "Closed Trades",
                "Daily Returns",
                "Deals",
                "Charts",
            ],
        )
        summary_headers = [
            cell.value for cell in next(exported_workbook["Summary"].iter_rows())
        ]
        self.assertIn("Sharpe Ratio", summary_headers)
        self.assertIn("Max Drawdown (%)", summary_headers)
        self.assertIn("Return on $100k (%)", summary_headers)
        return_column = summary_headers.index("Return on $100k (%)") + 1
        self.assertAlmostEqual(
            exported_workbook["Summary"].cell(2, return_column).value,
            0.1,
        )
        self.assertEqual(
            exported_workbook["Summary"].cell(2, return_column).number_format,
            '0.00"%"',
        )
        self.assertEqual(exported_workbook["Closed Trades"].max_row, 2)
        self.assertGreaterEqual(exported_workbook["Daily Returns"].max_row, 2)
        exported_workbook.close()


if __name__ == "__main__":
    unittest.main()
