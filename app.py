from __future__ import annotations

from datetime import timedelta
from io import BytesIO
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from plotly import express as px
from plotly import graph_objects as go

INITIAL_EQUITY = 100000.0
TRADING_DAYS_PER_YEAR = 252
EPSILON = 1e-9

DEAL_COLUMNS = [
    "Time",
    "Deal",
    "Position",
    "Order",
    "Symbol",
    "Type",
    "Direction",
    "Volume",
    "Profit",
    "Commission",
    "Fee",
    "Swap",
    "NetPnl",
]


SUMMARY_KEYS = {
    "total net profit": "Total Net Profit",
    "profit factor": "Profit Factor",
    "expected payoff": "Expected Payoff",
    "recovery factor": "Recovery Factor",
    "balance drawdown absolute": "Balance Drawdown Absolute",
    "balance drawdown maximal": "Balance Drawdown Maximal",
    "balance drawdown relative": "Balance Drawdown Relative",
    "sharpe ratio": "Sharpe Ratio",
    "total trades": "Total Trades",
    "profit trades (% of total)": "Profit Trades Percent",
}


NUMERIC_METRICS = {
    "Total Net Profit",
    "Profit Factor",
    "Expected Payoff",
    "Recovery Factor",
    "Balance Drawdown Absolute",
    "Sharpe Ratio",
    "Total Trades",
}


def parse_max_drawdown_value(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    first = text.split("(")[0].strip().replace(" ", "")
    try:
        return float(first)
    except ValueError:
        return None


def parse_percent_from_trade_stat(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value)
    if "(" not in text or "%" not in text:
        return None
    inside = text.split("(", 1)[1].split("%", 1)[0].strip()
    try:
        return float(inside)
    except ValueError:
        return None


def normalize_label(value: Any) -> str:
    return str(value).strip().lower().replace(":", "")


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    if "(" in text and ")" in text:
        text = text.split("(")[0].strip()

    text = text.replace(" ", "")
    try:
        return float(text)
    except ValueError:
        return None


def to_identifier(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def find_best_value_in_row(row: tuple[Any, ...], label_index: int) -> Any:
    for value in row[label_index + 1 :]:
        if value not in (None, ""):
            return value
    return None


def extract_summary_metrics(ws) -> dict[str, Any]:
    metrics: dict[str, Any] = {
        "Total Net Profit": None,
        "Profit Factor": None,
        "Expected Payoff": None,
        "Recovery Factor": None,
        "Balance Drawdown Absolute": None,
        "Balance Drawdown Maximal": None,
        "Balance Drawdown Relative": None,
        "Sharpe Ratio": None,
        "Balance Drawdown Maximal Value": None,
        "Total Trades": None,
        "Profit Trades Percent": None,
    }

    for row in ws.iter_rows(values_only=True):
        row_values = list(row)
        labels = [normalize_label(v) for v in row_values if v not in (None, "")]
        if not labels:
            continue

        for idx, raw in enumerate(row_values):
            if raw in (None, ""):
                continue
            label = normalize_label(raw)
            if label in SUMMARY_KEYS:
                metric_name = SUMMARY_KEYS[label]
                row_value = find_best_value_in_row(tuple(row_values), idx)
                if metric_name in NUMERIC_METRICS:
                    metrics[metric_name] = to_float(row_value)
                else:
                    metrics[metric_name] = row_value

                if metric_name == "Balance Drawdown Maximal":
                    metrics["Balance Drawdown Maximal Value"] = parse_max_drawdown_value(
                        row_value
                    )
                if metric_name == "Profit Trades Percent":
                    metrics["Profit Trades Percent"] = parse_percent_from_trade_stat(
                        row_value
                    )

    return metrics


def find_deals_header(ws) -> tuple[int, dict[str, int]]:
    max_col = ws.max_column
    for r in range(1, ws.max_row + 1):
        first_cell = ws.cell(r, 1).value
        if normalize_label(first_cell) == "deals":
            header_row = r + 1
            headers = [ws.cell(header_row, c).value for c in range(1, max_col + 1)]
            header_map = {
                normalize_label(name): idx + 1
                for idx, name in enumerate(headers)
                if name not in (None, "")
            }
            return header_row, header_map
    raise ValueError("Deals section not found")


def extract_deals_dataframe(ws) -> pd.DataFrame:
    header_row, header_map = find_deals_header(ws)
    required = ["time", "type", "direction", "profit"]
    for key in required:
        if key not in header_map:
            raise ValueError(f"Deals table is missing required column: {key}")

    rows: list[dict[str, Any]] = []
    r = header_row + 1
    valid_directions = {"in", "out", "inout", "out by"}

    while r <= ws.max_row:
        first_cell = ws.cell(r, 1).value
        if first_cell in (None, ""):
            # allow sparse blank rows after deals and before summary
            if normalize_label(ws.cell(r + 1, 1).value) in {
                "total net profit",
                "profit factor",
                "recovery factor",
                "balance drawdown",
            }:
                break
            r += 1
            continue

        first_label = normalize_label(first_cell)
        if first_label in {
            "total net profit",
            "profit factor",
            "recovery factor",
            "balance drawdown",
            "open time",
            "orders",
            "positions",
            "deals",
        }:
            break

        row_type = ws.cell(r, header_map["type"]).value
        direction = ws.cell(r, header_map["direction"]).value
        row_time = ws.cell(r, header_map["time"]).value
        row_profit = ws.cell(r, header_map["profit"]).value
        row_commission = (
            ws.cell(r, header_map["commission"]).value if "commission" in header_map else 0.0
        )
        row_fee = ws.cell(r, header_map["fee"]).value if "fee" in header_map else 0.0
        row_swap = ws.cell(r, header_map["swap"]).value if "swap" in header_map else 0.0
        row_deal = ws.cell(r, header_map["deal"]).value if "deal" in header_map else None
        row_position = (
            ws.cell(r, header_map["position"]).value if "position" in header_map else None
        )
        row_order = ws.cell(r, header_map["order"]).value if "order" in header_map else None
        row_symbol = ws.cell(r, header_map["symbol"]).value if "symbol" in header_map else None
        row_volume = ws.cell(r, header_map["volume"]).value if "volume" in header_map else 0.0

        row_time_parsed = pd.to_datetime(row_time, errors="coerce")
        if pd.isna(row_time_parsed):
            r += 1
            continue

        # Ignore rows from subsequent non-deal sections (Open positions/orders)
        # that can appear after the Deals section in some report exports.
        if to_identifier(row_deal) is None:
            r += 1
            continue

        row_type_norm = str(row_type).lower() if row_type is not None else ""
        direction_norm = str(direction).lower() if direction is not None else ""
        profit_value = to_float(row_profit) or 0.0
        commission_value = to_float(row_commission) or 0.0
        fee_value = to_float(row_fee) or 0.0
        swap_value = to_float(row_swap) or 0.0
        net_pnl = profit_value + commission_value + fee_value + swap_value

        if row_type_norm not in {"balance", "credit"} and direction_norm not in valid_directions:
            r += 1
            continue

        rows.append(
            {
                "Time": row_time_parsed,
                "Deal": to_identifier(row_deal),
                "Position": to_identifier(row_position),
                "Order": to_identifier(row_order),
                "Symbol": str(row_symbol).strip() if row_symbol not in (None, "") else None,
                "Type": row_type_norm,
                "Direction": direction_norm,
                "Volume": to_float(row_volume) or 0.0,
                "Profit": profit_value,
                "Commission": commission_value,
                "Fee": fee_value,
                "Swap": swap_value,
                "NetPnl": net_pnl,
            }
        )
        r += 1

    if not rows:
        return pd.DataFrame(columns=DEAL_COLUMNS)

    deals = pd.DataFrame(rows, columns=DEAL_COLUMNS)
    deals = deals.dropna(subset=["Time"]).sort_values("Time").reset_index(drop=True)
    return deals


def extract_realized_deals(deals: pd.DataFrame) -> pd.DataFrame:
    """Return one fully netted result per completed position/trade.

    MetaTrader's Deals section contains execution cash flows, not one row per
    trade. Entry commissions therefore must not be counted as separate losing
    trades. When position identifiers are available, all executions belonging
    to a closed position are grouped. Otherwise, closing executions form the
    trades and unmatched entry costs are allocated by closing volume so that
    trade P&L reconciles exactly to total trading cash flow.
    """
    if deals.empty:
        return pd.DataFrame(columns=["Time", "NetPnl", "Volume", "Position"])

    closing_directions = {"out", "inout", "out by"}
    non_trade_types = {"balance", "credit"}
    cashflows = deals[
        (~deals["Type"].isin(non_trade_types))
        & deals["Direction"].isin({"in", *closing_directions})
    ].copy()
    if cashflows.empty:
        return pd.DataFrame(columns=["Time", "NetPnl", "Volume", "Position"])

    closing = cashflows[cashflows["Direction"].isin(closing_directions)].copy()
    if closing.empty:
        return pd.DataFrame(columns=["Time", "NetPnl", "Volume", "Position"])

    rows: list[dict[str, Any]] = []
    used_indexes: set[int] = set()
    has_positions = (
        "Position" in cashflows.columns
        and closing["Position"].notna().any()
    )

    if has_positions:
        for position, closing_group in closing.dropna(subset=["Position"]).groupby(
            "Position", sort=False
        ):
            position_rows = cashflows[cashflows["Position"] == position]
            used_indexes.update(position_rows.index.tolist())
            rows.append(
                {
                    "Time": closing_group["Time"].max(),
                    "NetPnl": float(position_rows["NetPnl"].sum()),
                    "Volume": float(closing_group["Volume"].abs().sum()),
                    "Position": position,
                }
            )

    remaining_closes = closing[~closing.index.isin(used_indexes)]
    for idx, row in remaining_closes.iterrows():
        used_indexes.add(idx)
        rows.append(
            {
                "Time": row["Time"],
                "NetPnl": float(row["NetPnl"]),
                "Volume": abs(float(row.get("Volume", 0.0))),
                "Position": row.get("Position"),
            }
        )

    realized = pd.DataFrame(rows).sort_values("Time").reset_index(drop=True)

    # Entry fees may not carry a position ID in some exports. Reconcile any
    # remaining trading cash flow to completed trades without inventing trades.
    unmatched_pnl = float(cashflows.loc[~cashflows.index.isin(used_indexes), "NetPnl"].sum())
    if abs(unmatched_pnl) > EPSILON and not realized.empty:
        weights = realized["Volume"].abs().astype(float)
        if float(weights.sum()) <= EPSILON:
            weights = pd.Series(1.0, index=realized.index)
        realized["NetPnl"] += unmatched_pnl * (weights / float(weights.sum()))

    return realized


def build_daily_returns(realized_deals: pd.DataFrame) -> pd.DataFrame:
    """Build fixed-balance daily returns, including zero-P&L business days."""
    if realized_deals.empty:
        return pd.DataFrame(columns=["Date", "NetPnl", "Return"])

    daily_pnl = (
        realized_deals.assign(Date=pd.to_datetime(realized_deals["Time"]).dt.normalize())
        .groupby("Date")["NetPnl"]
        .sum()
        .sort_index()
    )
    business_days = pd.bdate_range(daily_pnl.index.min(), daily_pnl.index.max())
    full_index = business_days.union(daily_pnl.index).sort_values()
    daily_pnl = daily_pnl.reindex(full_index, fill_value=0.0)

    return pd.DataFrame(
        {
            "Date": daily_pnl.index,
            "NetPnl": daily_pnl.to_numpy(dtype=float),
            "Return": daily_pnl.to_numpy(dtype=float) / INITIAL_EQUITY,
        }
    )


def format_timedelta(value: timedelta | None) -> str:
    if value is None:
        return "N/A"

    total_seconds = int(max(value.total_seconds(), 0))
    total_hours = total_seconds // 3600
    days, hours = divmod(total_hours, 24)
    return f"{days}d {hours}h"


def timedelta_to_days(value: timedelta | None) -> float | None:
    if value is None:
        return None
    return value.total_seconds() / 86400.0


def max_drawdown_duration(drawdown_curve: pd.DataFrame) -> timedelta:
    if drawdown_curve.empty:
        return timedelta(0)

    curve = drawdown_curve.sort_values("Time").reset_index(drop=True)
    max_duration = timedelta(0)
    peak_time = pd.to_datetime(curve.iloc[0]["Time"])
    peak_equity = float(curve.iloc[0]["Equity"])
    underwater = False

    for _, row in curve.iloc[1:].iterrows():
        equity = float(row["Equity"])
        t = pd.to_datetime(row["Time"])
        if equity + EPSILON >= peak_equity:
            if underwater:
                max_duration = max(max_duration, t - peak_time)
                underwater = False
            peak_equity = max(peak_equity, equity)
            peak_time = t
        else:
            underwater = True

    if underwater:
        max_duration = max(
            max_duration,
            pd.to_datetime(curve.iloc[-1]["Time"]) - peak_time,
        )

    return max_duration


def compute_risk_metrics(
    realized_deals: pd.DataFrame,
    drawdown_curve: pd.DataFrame,
    daily_returns: pd.DataFrame | None = None,
    total_costs: float = 0.0,
) -> dict[str, Any]:
    daily = daily_returns if daily_returns is not None else build_daily_returns(realized_deals)
    if realized_deals.empty:
        return {
            "Total Net Profit": 0.0,
            "Gross Profit": 0.0,
            "Gross Loss": 0.0,
            "Costs / Swap": total_costs,
            "Profit Factor": None,
            "Total Trades": 0,
            "Win Rate": None,
            "Sharpe Ratio": None,
            "Sortino Ratio": None,
            "Time to Recovery": format_timedelta(timedelta(0)),
            "Time to Recovery Days": 0.0,
            "Largest Single Loss": 0.0,
            "Balance Drawdown Maximal Value": 0.0,
            "Max Drawdown (%)": 0.0,
            "Recovery Factor": None,
            "Calmar Ratio": None,
            "Net Profit / Max DD": None,
            "Expectancy per Trade": None,
            "Average Trade": None,
            "Average Win": None,
            "Average Loss": None,
            "Best Trade": None,
            "Worst Trade": None,
            "Best Day": None,
            "Worst Day": None,
            "Return on $100k": 0.0,
            "Annualized Return": None,
        }

    profits = realized_deals["NetPnl"].astype(float)
    trade_count = len(profits)
    wins = profits[profits > EPSILON]
    losses = profits[profits < -EPSILON]
    net_profit = float(profits.sum())
    gross_profit = float(wins.sum())
    gross_loss = float(losses.sum())
    win_rate = (len(wins) / trade_count) * 100.0 if trade_count else None
    profit_factor = gross_profit / abs(gross_loss) if gross_loss < -EPSILON else None
    expectancy = net_profit / trade_count if trade_count else None

    returns = daily["Return"].astype(float) if not daily.empty else pd.Series(dtype=float)
    mean_ret = float(returns.mean()) if not returns.empty else 0.0
    std_ret = float(returns.std(ddof=1)) if len(returns) > 1 else 0.0
    sharpe = (
        (mean_ret / std_ret) * (TRADING_DAYS_PER_YEAR**0.5)
        if std_ret > EPSILON
        else None
    )
    downside_returns = returns.clip(upper=0.0)
    downside_deviation = (
        float((downside_returns.pow(2).mean()) ** 0.5)
        if not downside_returns.empty
        else 0.0
    )
    sortino = (
        (mean_ret / downside_deviation) * (TRADING_DAYS_PER_YEAR**0.5)
        if downside_deviation > EPSILON
        else None
    )
    annualized_return = mean_ret * TRADING_DAYS_PER_YEAR if not returns.empty else None

    max_dd = float(drawdown_curve["Drawdown"].max()) if not drawdown_curve.empty else 0.0
    max_dd_pct = (
        float((drawdown_curve["Drawdown"] / drawdown_curve["Peak"]).max())
        if not drawdown_curve.empty
        else 0.0
    )
    recovery_factor = net_profit / max_dd if max_dd > EPSILON else None
    calmar = (
        annualized_return / max_dd_pct
        if annualized_return is not None and max_dd_pct > EPSILON
        else None
    )
    recovery_duration = max_drawdown_duration(drawdown_curve)
    daily_pnl = daily["NetPnl"].astype(float) if not daily.empty else pd.Series(dtype=float)
    trading_day_pnl = daily_pnl[daily_pnl.abs() > EPSILON]

    return {
        "Total Net Profit": net_profit,
        "Gross Profit": gross_profit,
        "Gross Loss": gross_loss,
        "Costs / Swap": total_costs,
        "Profit Factor": profit_factor,
        "Total Trades": trade_count,
        "Win Rate": win_rate,
        "Sharpe Ratio": sharpe,
        "Sortino Ratio": sortino,
        "Time to Recovery": format_timedelta(recovery_duration),
        "Time to Recovery Days": timedelta_to_days(recovery_duration),
        "Largest Single Loss": float(losses.min()) if not losses.empty else 0.0,
        "Balance Drawdown Maximal Value": max_dd,
        "Max Drawdown (%)": max_dd_pct,
        "Recovery Factor": recovery_factor,
        "Calmar Ratio": calmar,
        "Net Profit / Max DD": recovery_factor,
        "Expectancy per Trade": expectancy,
        "Average Trade": expectancy,
        "Average Win": float(wins.mean()) if not wins.empty else None,
        "Average Loss": float(losses.mean()) if not losses.empty else None,
        "Best Trade": float(profits.max()),
        "Worst Trade": float(profits.min()),
        "Best Day": float(trading_day_pnl.max()) if not trading_day_pnl.empty else None,
        "Worst Day": float(trading_day_pnl.min()) if not trading_day_pnl.empty else None,
        "Return on $100k": net_profit / INITIAL_EQUITY,
        "Annualized Return": annualized_return,
    }


def build_equity_curve(deals: pd.DataFrame) -> pd.DataFrame:
    if deals.empty:
        return pd.DataFrame(columns=["Time", "Profit", "Equity"])

    realized = extract_realized_deals(deals)

    if realized.empty:
        start_time = deals["Time"].min()
        return pd.DataFrame(
            [
                {
                    "Time": start_time,
                    "Profit": 0.0,
                    "Equity": INITIAL_EQUITY,
                }
            ]
        )

    # Prevent false drawdown spikes from arbitrary row order when multiple deals
    # share the same timestamp by aggregating realized PnL per timestamp first.
    realized_by_time = (
        realized.groupby("Time", sort=True, as_index=False)
        .agg({"NetPnl": "sum"})
    )
    realized_by_time["Equity"] = INITIAL_EQUITY + realized_by_time["NetPnl"].cumsum()

    anchor_time = realized_by_time["Time"].min() - timedelta(seconds=1)
    anchor = pd.DataFrame(
        [{"Time": anchor_time, "Profit": 0.0, "Equity": INITIAL_EQUITY}]
    )

    realized_curve = realized_by_time[["Time", "NetPnl", "Equity"]].copy()
    realized_curve.columns = ["Time", "Profit", "Equity"]

    return pd.concat(
        [anchor, realized_curve],
        ignore_index=True,
    )


def parse_report(uploaded_file) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True, read_only=False)
    ws = workbook[workbook.sheetnames[0]]

    reported_metrics = extract_summary_metrics(ws)
    deals = extract_deals_dataframe(ws)
    realized_deals = extract_realized_deals(deals)
    equity = build_equity_curve(deals)
    drawdown = build_drawdown_curve(equity)
    daily_returns = build_daily_returns(realized_deals)
    trade_deals = deals[~deals["Type"].isin({"balance", "credit"})]
    total_costs = float(
        trade_deals[["Commission", "Fee", "Swap"]].sum().sum()
    )
    metrics = compute_risk_metrics(
        realized_deals,
        drawdown,
        daily_returns,
        total_costs=total_costs,
    )

    workbook.close()
    return {
        "name": uploaded_file.name,
        "metrics": metrics,
        "reported_metrics": reported_metrics,
        "deals": deals,
        "realized_deals": realized_deals,
        "daily_returns": daily_returns,
        "equity": equity,
        "drawdown": drawdown,
    }


def build_summary_table(reports: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for report in reports:
        row = {"Report": report["name"]}
        row.update(report["metrics"])
        rows.append(row)
    return pd.DataFrame(rows)


def build_display_summary_table(summary_df: pd.DataFrame) -> pd.DataFrame:
    preferred_order = [
        "Report",
        "Total Net Profit",
        "Gross Profit",
        "Gross Loss",
        "Costs / Swap",
        "Balance Drawdown Maximal Value",
        "Max Drawdown (%)",
        "Return on $100k",
        "Annualized Return",
        "Win Rate",
        "Profit Factor",
        "Sharpe Ratio",
        "Sortino Ratio",
        "Calmar Ratio",
        "Time to Recovery",
        "Recovery Factor",
        "Average Win",
        "Average Loss",
        "Best Trade",
        "Best Day",
        "Worst Day",
        "Expectancy per Trade",
        "Largest Single Loss",
        "Total Trades",
    ]
    present_cols = [c for c in preferred_order if c in summary_df.columns]
    display_summary = summary_df[present_cols].copy()
    for percentage_col in [
        "Max Drawdown (%)",
        "Return on $100k",
        "Annualized Return",
    ]:
        if percentage_col in display_summary.columns:
            display_summary[percentage_col] *= 100.0
    display_summary = display_summary.rename(
        columns={
            "Balance Drawdown Maximal Value": "Max Drawdown ($)",
            "Return on $100k": "Return on $100k (%)",
            "Annualized Return": "Annualized Return (%)",
            "Win Rate": "Win Rate (%)",
        }
    )
    return display_summary


def build_equity_overlay_table(reports: list[dict[str, Any]]) -> pd.DataFrame:
    merged: pd.DataFrame | None = None
    for report in reports:
        curve = report["equity"][["Time", "Equity"]].rename(
            columns={"Equity": report["name"]}
        )
        if merged is None:
            merged = curve
        else:
            merged = merged.merge(curve, on="Time", how="outer")

    if merged is None:
        return pd.DataFrame(columns=["Time"])

    merged = merged.sort_values("Time")
    report_columns = [column for column in merged.columns if column != "Time"]
    merged[report_columns] = merged[report_columns].ffill().fillna(INITIAL_EQUITY)
    return merged


def build_drawdown_curve(equity_curve: pd.DataFrame) -> pd.DataFrame:
    if equity_curve.empty:
        return pd.DataFrame(columns=["Time", "Equity", "Peak", "Drawdown"])

    curve = equity_curve[["Time", "Equity"]].copy().sort_values("Time")
    curve["Peak"] = curve["Equity"].cummax()
    curve["Drawdown"] = (curve["Peak"] - curve["Equity"]).clip(lower=0)
    return curve[["Time", "Equity", "Peak", "Drawdown"]]


def make_download_workbook(reports: list[dict[str, Any]]) -> bytes:
    summary_df = build_summary_table(reports)
    display_summary = build_display_summary_table(summary_df)
    equity_overlay = build_equity_overlay_table(reports)

    all_deals = []
    all_realized = []
    all_daily_returns = []
    for report in reports:
        deals = report["deals"].copy()
        deals.insert(0, "Report", report["name"])
        all_deals.append(deals)
        realized = report["realized_deals"].copy()
        realized.insert(0, "Report", report["name"])
        all_realized.append(realized)
        daily_returns = report["daily_returns"].copy()
        daily_returns.insert(0, "Report", report["name"])
        all_daily_returns.append(daily_returns)
    deals_df = pd.concat(all_deals, ignore_index=True) if all_deals else pd.DataFrame()
    realized_df = (
        pd.concat(all_realized, ignore_index=True) if all_realized else pd.DataFrame()
    )
    daily_returns_df = (
        pd.concat(all_daily_returns, ignore_index=True)
        if all_daily_returns
        else pd.DataFrame()
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter", datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
        display_summary.to_excel(writer, sheet_name="Summary", index=False)
        equity_overlay.to_excel(writer, sheet_name="Equity Overlay", index=False)
        realized_df.to_excel(writer, sheet_name="Closed Trades", index=False)
        daily_returns_df.to_excel(writer, sheet_name="Daily Returns", index=False)
        deals_df.to_excel(writer, sheet_name="Deals", index=False)

        workbook = writer.book
        summary_sheet = writer.sheets["Summary"]
        equity_sheet = writer.sheets["Equity Overlay"]
        charts_sheet = workbook.add_worksheet("Charts")

        currency_fmt = workbook.add_format({"num_format": "$#,##0.00"})
        ratio_fmt = workbook.add_format({"num_format": "0.000"})
        for col_name in [
            "Total Net Profit",
            "Max Drawdown ($)",
            "Largest Single Loss",
            "Expectancy per Trade",
            "Gross Profit",
            "Gross Loss",
            "Average Win",
            "Average Loss",
            "Best Trade",
            "Worst Trade",
            "Best Day",
            "Worst Day",
        ]:
            if col_name in display_summary.columns:
                col_idx = display_summary.columns.get_loc(col_name)
                summary_sheet.set_column(col_idx, col_idx, 22, currency_fmt)

        for col_name in [
            "Profit Factor",
            "Sharpe Ratio",
            "Sortino Ratio",
            "Calmar Ratio",
            "Recovery Factor",
        ]:
            if col_name in display_summary.columns:
                col_idx = display_summary.columns.get_loc(col_name)
                summary_sheet.set_column(col_idx, col_idx, 16, ratio_fmt)

        percentage_fmt = workbook.add_format({"num_format": '0.00"%"'})
        for col_name in [
            "Win Rate (%)",
            "Max Drawdown (%)",
            "Return on $100k (%)",
            "Annualized Return (%)",
        ]:
            if col_name in display_summary.columns:
                col_idx = display_summary.columns.get_loc(col_name)
                summary_sheet.set_column(col_idx, col_idx, 18, percentage_fmt)

        if "Time to Recovery" in display_summary.columns:
            col_idx = display_summary.columns.get_loc("Time to Recovery")
            summary_sheet.set_column(col_idx, col_idx, 16)

        equity_sheet.set_column(0, 0, 22)
        equity_sheet.set_column(1, max(1, len(equity_overlay.columns) - 1), 18, currency_fmt)

        row_count = len(display_summary)
        if row_count > 0 and "Total Net Profit" in display_summary.columns:
            chart_profit = workbook.add_chart({"type": "column"})
            metric_col = display_summary.columns.get_loc("Total Net Profit")
            chart_profit.add_series(
                {
                    "name": "Total Net Profit",
                    "categories": ["Summary", 1, 0, row_count, 0],
                    "values": ["Summary", 1, metric_col, row_count, metric_col],
                    "data_labels": {"value": True},
                }
            )
            chart_profit.set_title({"name": "Total Net Profit by Report"})
            chart_profit.set_y_axis({"name": "$", "num_format": "$#,##0"})
            chart_profit.set_legend({"none": True})
            charts_sheet.insert_chart("B2", chart_profit, {"x_scale": 1.2, "y_scale": 1.2})

        if row_count > 0 and "Max Drawdown ($)" in display_summary.columns:
            chart_drawdown = workbook.add_chart({"type": "column"})
            dd_col = display_summary.columns.get_loc("Max Drawdown ($)")
            chart_drawdown.add_series(
                {
                    "name": "Max Drawdown ($)",
                    "categories": ["Summary", 1, 0, row_count, 0],
                    "values": ["Summary", 1, dd_col, row_count, dd_col],
                    "data_labels": {"value": True},
                }
            )
            chart_drawdown.set_title({"name": "Max Drawdown by Report"})
            chart_drawdown.set_y_axis({"name": "$", "num_format": "$#,##0"})
            chart_drawdown.set_legend({"none": True})
            charts_sheet.insert_chart("B20", chart_drawdown, {"x_scale": 1.2, "y_scale": 1.2})

        if len(equity_overlay.columns) > 1 and len(equity_overlay) > 1:
            chart_equity = workbook.add_chart({"type": "line"})
            for col_idx in range(1, len(equity_overlay.columns)):
                chart_equity.add_series(
                    {
                        "name": ["Equity Overlay", 0, col_idx],
                        "categories": ["Equity Overlay", 1, 0, len(equity_overlay), 0],
                        "values": ["Equity Overlay", 1, col_idx, len(equity_overlay), col_idx],
                    }
                )
            chart_equity.set_title({"name": "Equity Curve Overlay (Start $100,000)"})
            chart_equity.set_y_axis({"name": "Equity ($)", "num_format": "$#,##0"})
            chart_equity.set_x_axis({"name": "Time"})
            charts_sheet.insert_chart("K2", chart_equity, {"x_scale": 1.5, "y_scale": 1.4})

    output.seek(0)
    return output.getvalue()


def render_dashboard(reports: list[dict[str, Any]]) -> None:
    summary_df = build_summary_table(reports)
    display_summary = build_display_summary_table(summary_df)

    # Keep the original table at the top for full report context.
    st.subheader("Key Metrics by Report")
    display_summary_view = display_summary.copy()
    display_summary_view.index = list(range(1, len(display_summary_view) + 1))
    display_summary_view.index.name = "Row"
    st.dataframe(display_summary_view, use_container_width=True)

    # 1. Equity Curve (highest priority)
    st.subheader("Equity Curve Overlay (Start $100,000)")
    curves = []
    for report in reports:
        curve = report["equity"].copy()
        curve["Report"] = report["name"]
        curves.append(curve)

    if curves:
        long_equity = pd.concat(curves, ignore_index=True)
        fig_line = px.line(
            long_equity,
            x="Time",
            y="Equity",
            color="Report",
            title="Equity Curves Across Uploaded Reports",
            labels={"Equity": "Equity ($)"},
        )
        st.plotly_chart(fig_line, use_container_width=True)

    # 2 & 3. PnL and Drawdown in one chart for direct payoff vs pain comparison.
    if {"Total Net Profit", "Balance Drawdown Maximal Value"}.issubset(summary_df.columns):
        st.subheader("Profit vs Max Drawdown")
        compare_df = summary_df[
            ["Report", "Total Net Profit", "Balance Drawdown Maximal Value"]
        ].copy()
        compare_melted = compare_df.melt(
            id_vars=["Report"],
            value_vars=["Total Net Profit", "Balance Drawdown Maximal Value"],
            var_name="Metric",
            value_name="Value",
        )
        compare_melted["Metric"] = compare_melted["Metric"].replace(
            {"Balance Drawdown Maximal Value": "Balance Drawdown Maximal ($)"}
        )
        fig_profit_dd = px.bar(
            compare_melted,
            x="Report",
            y="Value",
            color="Metric",
            barmode="group",
            title="Net Profit and Max Drawdown (Absolute)",
            labels={"Value": "$"},
        )
        st.plotly_chart(fig_profit_dd, use_container_width=True)

    # 4, 5, 6. Show the actual values; normalizing ratios against the batch can
    # hide negative Sharpe values and is not a valid performance calculation.
    quality_cols = [c for c in ["Win Rate", "Profit Factor", "Sharpe Ratio"] if c in summary_df.columns]
    if quality_cols:
        st.subheader("Trade Quality and Risk-Adjusted Performance")
        st.caption(
            "Raw calculated values are shown on independent axes so unlike metrics "
            "are not converted into a synthetic score."
        )
        quality_df = summary_df[["Report", *quality_cols]].copy()
        quality_melted = quality_df.melt(
            id_vars=["Report"],
            value_vars=quality_cols,
            var_name="Metric",
            value_name="Value",
        )
        fig_quality = px.bar(
            quality_melted,
            x="Report",
            y="Value",
            facet_col="Metric",
            facet_col_wrap=3,
            title="Trade Quality Metrics (Raw Values)",
        )
        fig_quality.update_yaxes(matches=None)
        fig_quality.for_each_annotation(lambda annotation: annotation.update(text=annotation.text.split("=")[-1]))
        st.plotly_chart(fig_quality, use_container_width=True)

    # 7. Time to recovery shown as days/hours labels and numeric day bars.
    if {"Time to Recovery", "Time to Recovery Days"}.issubset(summary_df.columns):
        st.subheader("Time to Recovery")
        ttr_df = summary_df[["Report", "Time to Recovery", "Time to Recovery Days"]].copy()
        ttr_max = float(pd.to_numeric(ttr_df["Time to Recovery Days"], errors="coerce").max() or 0.0)
        fig_ttr = go.Figure(
            data=[
                go.Bar(
                    x=ttr_df["Report"],
                    y=ttr_df["Time to Recovery Days"],
                    text=ttr_df["Time to Recovery"],
                    textposition="outside",
                    cliponaxis=False,
                    name="Recovery Duration",
                )
            ]
        )
        fig_ttr.update_layout(
            title="Recovery Duration by Report",
            xaxis_title="Report",
            yaxis_title="Days",
            margin={"t": 80, "r": 20, "b": 60, "l": 50},
        )
        if ttr_max > 0:
            fig_ttr.update_yaxes(range=[0, ttr_max * 1.20])
        st.plotly_chart(fig_ttr, use_container_width=True)

    # 8. Secondary metrics: efficiency and expectancy with explicit economic meaning.
    secondary_cols = [c for c in ["Expectancy per Trade", "Largest Single Loss", "Sortino Ratio", "Recovery Factor"] if c in summary_df.columns]
    if secondary_cols:
        st.subheader("Secondary Metrics")
        sec_df = summary_df[["Report", *secondary_cols]].copy()
        sec_melted = sec_df.melt(
            id_vars=["Report"],
            value_vars=secondary_cols,
            var_name="Metric",
            value_name="Value",
        )
        fig_secondary = px.bar(
            sec_melted,
            x="Report",
            y="Value",
            facet_col="Metric",
            facet_col_wrap=2,
            title="Efficiency, Expectancy, and Tail-Risk Metrics",
        )
        fig_secondary.update_yaxes(matches=None)
        fig_secondary.for_each_annotation(
            lambda annotation: annotation.update(text=annotation.text.split("=")[-1])
        )
        st.plotly_chart(fig_secondary, use_container_width=True)


def main() -> None:
    st.set_page_config(page_title="Arxem Report Visualizer", layout="wide")
    st.title("Arxem Multi-Report Visualizer")

    uploaded_files = st.file_uploader(
        "Upload Excel files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Upload one or more MT trade history reports in .xlsx format.",
    )

    if not uploaded_files:
        st.info("Upload Excel files to generate analytics and charts.")
        return

    reports = []
    parsing_errors = []
    for file in uploaded_files:
        try:
            reports.append(parse_report(file))
        except Exception as exc:
            parsing_errors.append(f"{file.name}: {exc}")

    if parsing_errors:
        st.error("Some files could not be parsed:")
        for error in parsing_errors:
            st.write(f"- {error}")

    if not reports:
        st.warning("No valid reports were parsed.")
        return

    render_dashboard(reports)

    consolidated = make_download_workbook(reports)
    st.download_button(
        "Download Consolidate Report",
        data=consolidated,
        file_name="consolidated_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )


if __name__ == "__main__":
    main()
